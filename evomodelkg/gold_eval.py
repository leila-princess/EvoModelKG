from __future__ import annotations

import ast
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
import sys

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
PARENT = HERE.parent
if str(PARENT) not in sys.path:
    sys.path.insert(0, str(PARENT))

try:
    from evomodelkg.metrics import group_aware_values_match
except Exception:
    try:
        from evomodelkg.comparison.readme_structured_compare import values_match as _legacy_values_match
    except Exception:
        def group_aware_values_match(readme_val: Any, struct_val: Any, attr: str | None = None) -> bool:
            return False
    else:
        def group_aware_values_match(readme_val: Any, struct_val: Any, attr: str | None = None) -> bool:
            return _legacy_values_match(readme_val, struct_val, attr=attr)


TEXT_RECALL_ONLY_FIELDS = {
    "direct_use",
    "out_of_scope_use",
    "risks_and_biases",
}

DEFAULT_IGNORE_FIELDS = {
    "selection_rank",
    "task_group",
    "downloads",
    "readme_chars",
    "readme_file",
    "prelabel_status",
    "prelabel_error",
}

META_FIELDS = [
    "selection_rank",
    "model_id",
    "task_group",
    "downloads",
    "readme_chars",
    "readme_file",
    "prelabel_status",
    "prelabel_error",
]


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return False


def norm_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).lower()
    text = "".join(ch for ch in text if unicodedata.category(ch) not in {"Cf", "Cc"})
    text = text.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def compact_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm_text(value))


PIPELINE_TAG_ALIASES = {
    "image text to text": "image to text",
    "image to text": "image to text",
    "visual question answering": "image to text",
    "any to any": "image text to text",
    "text ranking": "text classification",
    "reranking": "text classification",
    "reranker": "text classification",
    "sentence similarity": "feature extraction",
    "text embeddings inference": "feature extraction",
    "text2text generation": "text generation",
    "text to text generation": "text generation",
}


LIBRARY_NAME_ALIASES = {
    "sentence transformers": "sentence-transformers",
    "sentence transformers library": "sentence-transformers",
    "transformers": "transformers",
    "huggingface transformers": "transformers",
    "pytorch": "torch",
    "torch": "torch",
    "autogptq": "auto-gptq",
    "auto gptq": "auto-gptq",
}


def canonical_label(value: Any, aliases: dict[str, str]) -> str:
    text = norm_text(value)
    return aliases.get(text, text)


def normalize_bibtex(value: Any) -> dict[str, str] | None:
    if is_empty(value):
        return None
    text = unicodedata.normalize("NFKC", str(value))
    text = re.sub(r"%.*", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text or "@" not in text:
        return None
    kind_key = re.search(r"@\s*([a-zA-Z]+)\s*\{\s*([^,\s]+)", text)
    fields: dict[str, str] = {}
    if kind_key:
        fields["entry_type"] = kind_key.group(1).lower()
        fields["key"] = norm_text(kind_key.group(2))
    for name in ("title", "author", "year", "eprint", "archiveprefix", "primaryclass", "url", "howpublished", "note"):
        pattern = rf"{name}\s*=\s*(?:\{{([^{{}}]*(?:\{{[^{{}}]*\}}[^{{}}]*)*)\}}|\"([^\"]*)\"|([^,}}]+))"
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if not m:
            continue
        raw = next((g for g in m.groups() if g is not None), "")
        raw = raw.replace(r"\url", "")
        raw = raw.strip(" {}\"',")
        fields[name.lower()] = compact_text(raw) if name != "year" else norm_text(raw)
    return fields or None


def bibtex_match(pred: Any, gold: Any) -> bool:
    a = normalize_bibtex(pred)
    b = normalize_bibtex(gold)
    if not a or not b:
        return False
    shared_keys = set(a) & set(b)
    strong_keys = {"key", "title", "author", "year", "eprint", "url"}
    shared_strong = shared_keys & strong_keys
    if not shared_strong:
        return compact_text(pred) == compact_text(gold)
    matches = sum(1 for key in shared_strong if a.get(key) == b.get(key))
    if {"title", "year"} <= shared_strong and a.get("title") == b.get("title") and a.get("year") == b.get("year"):
        return True
    return matches / max(1, len(shared_strong)) >= 0.75


def parse_maybe_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, set):
        return sorted(value)
    if isinstance(value, dict):
        return list(value.values())
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, (list, tuple, set)):
                return list(parsed)
        except Exception:
            pass
    return [text]


def scalar_tokens(value: Any, *, attr: str | None = None) -> set[str]:
    if is_empty(value):
        return set()
    if isinstance(value, bool):
        return {"true" if value else "false"}
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return {str(int(value))}
        return {str(value)}
    if isinstance(value, dict):
        out: set[str] = set()
        for item in value.values():
            out |= scalar_tokens(item, attr=attr)
        return out
    items = parse_maybe_list(value)
    if len(items) > 1:
        out: set[str] = set()
        for item in items:
            out |= scalar_tokens(item, attr=attr)
        return out
    text = str(items[0] if items else value).strip()
    if attr in {"training_datasets", "evaluation_datasets", "model_file_formats", "languages", "language"}:
        pieces = [p.strip(" '\"\t\r\n") for p in re.split(r"[,;|\n]+", text)]
        pieces = [p for p in pieces if p]
        if len(pieces) > 1:
            return {norm_text(p) for p in pieces}
    return {norm_text(text)}


def set_jaccard_match(pred: Any, gold: Any, *, attr: str, threshold: float) -> bool:
    pred_items = scalar_tokens(pred, attr=attr)
    gold_items = scalar_tokens(gold, attr=attr)
    if not pred_items or not gold_items:
        return False
    return len(pred_items & gold_items) / len(pred_items | gold_items) >= threshold


def item_level_f1_match(pred: Any, gold: Any, *, attr: str, threshold: float) -> bool:
    pred_items = scalar_tokens(pred, attr=attr)
    gold_items = scalar_tokens(gold, attr=attr)
    if not pred_items or not gold_items:
        return False
    matched = len(pred_items & gold_items)
    if matched == 0:
        return False
    precision = matched / len(pred_items)
    recall = matched / len(gold_items)
    f1 = 2 * precision * recall / (precision + recall)
    return f1 >= threshold


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = unicodedata.normalize("NFKC", str(value)).lower().replace(",", "").strip()
    if not text:
        return None
    m = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*([kmbt])", text)
    if m:
        return float(m.group(1)) * {"k": 1e3, "m": 1e6, "b": 1e9, "t": 1e12}[m.group(2)]
    m = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*(?:x|\*)\s*10\s*\^?\s*([+-]?\d+)", text)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    m = re.search(r"([+-]?\d+(?:\.\d+)?(?:e[+-]?\d+)?)", text)
    if not m:
        return None
    try:
        num = float(m.group(1))
    except Exception:
        return None
    suffix = text[m.end():].strip()[:1]
    if suffix in {"k", "m", "b", "t"}:
        num *= {"k": 1e3, "m": 1e6, "b": 1e9, "t": 1e12}[suffix]
    return num


def phrase_tokens(value: Any) -> set[str]:
    text = str(value or "")
    text = unicodedata.normalize("NFKC", text).lower()
    raw_parts = re.split(r"[,;|/\n、，；]+|\band\b|\bor\b", text)
    out: set[str] = set()
    stop = {"gpu", "gpus", "cpu", "cpus", "nvidia", "amd", "using", "with", "on", "the", "and"}
    for part in raw_parts:
        part = part.strip()
        if not part:
            continue
        compact = re.sub(r"[^a-z0-9]+", "", part)
        if compact and compact not in stop and len(compact) >= 2:
            out.add(compact)
        for tok in re.findall(r"[a-z]+[0-9]+[a-z0-9]*|[0-9]+[a-z]+[a-z0-9]*|[a-z]{2,}", part):
            if tok not in stop and len(tok) >= 2:
                out.add(tok)
    return out


def compute_infrastructure_match(pred: Any, gold: Any) -> bool:
    a = phrase_tokens(pred)
    b = phrase_tokens(gold)
    if not a or not b:
        return False
    if a & b:
        return True
    small, large = (a, b) if len(a) <= len(b) else (b, a)
    overlap = len(a & b) / max(1, len(small))
    if overlap >= 0.5:
        return True
    pred_norm = re.sub(r"[^a-z0-9]+", "", norm_text(pred))
    gold_norm = re.sub(r"[^a-z0-9]+", "", norm_text(gold))
    return bool(pred_norm and gold_norm and (pred_norm in gold_norm or gold_norm in pred_norm))


def values_match_for_gold(pred: Any, gold: Any, attr: str) -> bool:
    if is_empty(pred) or is_empty(gold):
        return False
    if attr == "citation_bibtex":
        return bibtex_match(pred, gold)
    if attr == "compute_infrastructure":
        return compute_infrastructure_match(pred, gold)
    if attr in {"languages", "model_file_formats"}:
        return set_jaccard_match(pred, gold, attr=attr, threshold=0.80)
    if attr in {"training_datasets", "evaluation_datasets"}:
        return item_level_f1_match(pred, gold, attr=attr, threshold=0.67)
    if attr in {
        "num_parameters",
        "context_length",
        "max_position_embeddings",
        "vocab_size",
        "hidden_size",
        "num_hidden_layers",
        "num_attention_heads",
    }:
        a = parse_number(pred)
        b = parse_number(gold)
        if a is not None and b is not None:
            rel_tol = 0.20 if attr == "num_parameters" else 0.02
            return abs(a - b) <= max(1e-9, rel_tol * max(abs(a), abs(b)))
    try:
        if group_aware_values_match(pred, gold, attr=attr):
            return True
    except Exception:
        pass
    a_tokens = scalar_tokens(pred, attr=attr)
    b_tokens = scalar_tokens(gold, attr=attr)
    if not a_tokens or not b_tokens:
        return False
    if attr in {"model_sub_types"}:
        return bool(a_tokens & b_tokens)
    if attr in {"cited_papers"}:
        return bool(a_tokens & b_tokens)
    if attr == "pipeline_tag":
        a_labels = {canonical_label(t, PIPELINE_TAG_ALIASES) for t in a_tokens}
        b_labels = {canonical_label(t, PIPELINE_TAG_ALIASES) for t in b_tokens}
        return bool(a_labels & b_labels)
    if attr == "library_name":
        a_labels = {canonical_label(t, LIBRARY_NAME_ALIASES) for t in a_tokens}
        b_labels = {canonical_label(t, LIBRARY_NAME_ALIASES) for t in b_tokens}
        return bool(a_labels & b_labels)
    if attr in {"model_id", "model_name", "author", "license", "license_name"}:
        return bool(a_tokens & b_tokens)
    a_join = " ".join(sorted(a_tokens))
    b_join = " ".join(sorted(b_tokens))
    return a_join in b_join or b_join in a_join or bool(a_tokens & b_tokens)


def load_gold_workbook(path: Path) -> tuple[list[str], dict[str, dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Gold_Final_Compact"] if "Gold_Final_Compact" in wb.sheetnames else wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        row = {str(headers[c - 1]): ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rid = str(row.get("model_id") or "").strip()
        if rid:
            rows[rid] = row
    return [str(h) for h in headers], rows


def load_extractions(path: Path) -> dict[str, dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "extractions" in payload:
        rows = payload["extractions"]
    else:
        rows = payload
    out: dict[str, dict[str, Any]] = {}
    for item in rows:
        rid = str(item.get("resource_id") or "").strip()
        extraction = item.get("extraction") if isinstance(item.get("extraction"), dict) else item
        if rid:
            out[rid] = extraction
    return out


def extraction_attr_map(extraction: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for row in extraction.get("attributes") or []:
        if not isinstance(row, dict):
            continue
        attr = str(row.get("attribute") or "").strip()
        if not attr or attr in out:
            continue
        out[attr] = row.get("value")
    return out


@dataclass
class EvalOptions:
    recall_only_fields: set[str]
    ignore_fields: set[str]


def evaluate_gold(
    gold_path: Path,
    extraction_path: Path,
    *,
    options: EvalOptions | None = None,
) -> dict[str, Any]:
    options = options or EvalOptions(TEXT_RECALL_ONLY_FIELDS, DEFAULT_IGNORE_FIELDS)
    headers, gold_rows = load_gold_workbook(gold_path)
    extractions = load_extractions(extraction_path)
    eval_fields = [
        h for h in headers
        if h not in options.ignore_fields and h not in {"model_id"} and not h.endswith("_id2")
    ]

    totals = Counter()
    per_field: dict[str, Counter] = defaultdict(Counter)
    mismatches: list[dict[str, Any]] = []

    for rid, gold in gold_rows.items():
        pred_map = extraction_attr_map(extractions.get(rid, {}))
        for field in eval_fields:
            gold_val = gold.get(field)
            pred_val = pred_map.get(field)
            g_has = not is_empty(gold_val)
            p_has = not is_empty(pred_val)
            recall_only = field in options.recall_only_fields

            if g_has:
                totals["gold_positive"] += 1
                per_field[field]["gold_positive"] += 1
                if p_has:
                    totals["recall_hit"] += 1
                    per_field[field]["recall_hit"] += 1
                else:
                    per_field[field]["fn"] += 1
                    totals["fn"] += 1

            if recall_only:
                continue

            if p_has:
                totals["pred_positive"] += 1
                per_field[field]["pred_positive"] += 1
                if g_has and values_match_for_gold(pred_val, gold_val, field):
                    totals["tp"] += 1
                    per_field[field]["tp"] += 1
                else:
                    totals["fp"] += 1
                    per_field[field]["fp"] += 1
                    mismatches.append({
                        "resource_id": rid,
                        "attribute": field,
                        "pred_value": pred_val,
                        "gold_value": gold_val,
                        "reason": "wrong_value" if g_has else "gold_empty_pred_present",
                    })
            elif g_has:
                # Missing a comparable gold value counts against recall/F1.
                pass

    comparable_gold_positive = sum(
        c["gold_positive"] for f, c in per_field.items()
        if f not in options.recall_only_fields
    )
    fn_for_f1 = comparable_gold_positive - totals["tp"]
    precision = totals["tp"] / totals["pred_positive"] if totals["pred_positive"] else 0.0
    recall = totals["recall_hit"] / totals["gold_positive"] if totals["gold_positive"] else 0.0
    comparable_recall = totals["tp"] / comparable_gold_positive if comparable_gold_positive else 0.0
    f1 = (
        2 * precision * comparable_recall / (precision + comparable_recall)
        if precision + comparable_recall > 0 else 0.0
    )

    per_field_out = {}
    for field in eval_fields:
        c = per_field[field]
        p = c["tp"] / c["pred_positive"] if c["pred_positive"] else None
        r_cov = c["recall_hit"] / c["gold_positive"] if c["gold_positive"] else None
        r_match = c["tp"] / c["gold_positive"] if c["gold_positive"] and field not in options.recall_only_fields else None
        f = (
            2 * p * r_match / (p + r_match)
            if p is not None and r_match is not None and p + r_match > 0 else None
        )
        per_field_out[field] = {
            "gold_positive": c["gold_positive"],
            "pred_positive": c["pred_positive"],
            "tp": c["tp"],
            "fp": c["fp"],
            "fn": c["fn"],
            "precision": p,
            "coverage_recall": r_cov,
            "matched_recall": r_match,
            "f1": f,
            "recall_only": field in options.recall_only_fields,
        }

    return {
        "summary": {
            "case_count": len(gold_rows),
            "field_count": len(eval_fields),
            "recall_only_fields": sorted(options.recall_only_fields),
            "gold_positive": totals["gold_positive"],
            "pred_positive_comparable": totals["pred_positive"],
            "tp": totals["tp"],
            "fp": totals["fp"],
            "fn_for_comparable_f1": fn_for_f1,
            "precision": precision,
            "recall_coverage_all_fields": recall,
            "matched_recall_comparable_fields": comparable_recall,
            "f1_comparable_fields": f1,
        },
        "per_field": per_field_out,
        "mismatches": mismatches,
    }


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--gold", required=True)
    parser.add_argument("--extractions", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    report = evaluate_gold(Path(args.gold), Path(args.extractions))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
